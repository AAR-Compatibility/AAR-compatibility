import ast
import os
import operator

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

engine = create_engine('postgresql://postgres:ikwillerencoderen156@localhost:5432/aar_comp_2')
BASE_DIR = os.path.dirname(__file__)
excel_path = os.path.join(BASE_DIR, "AAR_matrix_2.xlsx")

# Create the core lookup tables used by the import. The specifications table
# is created elsewhere in the SQL bootstrap script.
create_tables_sql = """
CREATE TABLE IF NOT EXISTS tankers(
    id SERIAL PRIMARY KEY,
    nation TEXT,
    type TEXT,
    model TEXT
);
CREATE TABLE IF NOT EXISTS receivers(
    id SERIAL PRIMARY KEY,
    nation TEXT,
    type TEXT,
    model TEXT
);
CREATE TABLE IF NOT EXISTS compatibility (
    id SERIAL PRIMARY KEY,
    tanker_id INTEGER NOT NULL REFERENCES tankers(id) ON DELETE CASCADE,
    receiver_id INTEGER NOT NULL REFERENCES receivers(id) ON DELETE CASCADE,
    UNIQUE (tanker_id, receiver_id)
);
"""

PAIR_COLUMNS = [
    'tanker_nation',
    'tanker_type',
    'tanker_model',
    'receiver_nation',
    'receiver_type',
    'receiver_model',
]

ALLOWED_ARITHMETIC_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}

ALLOWED_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def evaluate_simple_excel_formula(formula):
    # Only evaluate simple numeric formulas so formula-backed cells can still
    # be imported even when pandas does not return a computed value.
    expression = formula.lstrip("=").strip()
    node = ast.parse(expression, mode="eval")

    def walk(current):
        if isinstance(current, ast.Expression):
            return walk(current.body)
        if isinstance(current, ast.Constant) and isinstance(current.value, (int, float)):
            return float(current.value)
        if isinstance(current, ast.BinOp) and type(current.op) in ALLOWED_ARITHMETIC_OPERATORS:
            return ALLOWED_ARITHMETIC_OPERATORS[type(current.op)](
                walk(current.left),
                walk(current.right),
            )
        if isinstance(current, ast.UnaryOp) and type(current.op) in ALLOWED_UNARY_OPERATORS:
            return ALLOWED_UNARY_OPERATORS[type(current.op)](walk(current.operand))
        raise ValueError(f"Unsupported Excel formula: {formula}")

    return float(walk(node))


def load_formula_backed_column(workbook_path, sheet_name, column_name):
    # Read the raw worksheet so we can inspect literal cell contents, including
    # formulas that may not be pre-calculated in the workbook.
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    column_index = header_row.index(column_name)
    values = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        cell_value = row[column_index]
        if isinstance(cell_value, (int, float)):
            values.append(float(cell_value))
            continue
        if isinstance(cell_value, str):
            stripped = cell_value.strip()
            if not stripped:
                values.append(pd.NA)
                continue
            if stripped.startswith("="):
                values.append(round(evaluate_simple_excel_formula(stripped), 2))
                continue
            try:
                values.append(float(stripped.replace(",", "")))
                continue
            except ValueError:
                values.append(pd.NA)
                continue
        values.append(pd.NA)

    workbook.close()
    return pd.Series(values, dtype="Float64")


def clean_text_columns(frame, columns):
    cleaned = frame.copy()
    for column in columns:
        # Normalise whitespace to avoid failed joins caused by hidden spaces.
        cleaned[column] = cleaned[column].astype("string").str.strip()
    return cleaned


def build_master_rows(master_frame, specification_frame, source_columns):
    # Extend the master tanker/receiver lists with combinations that appear in
    # Specifications so later joins always have a matching base record.
    spec_rows = specification_frame[source_columns].rename(
        columns={
            source_columns[0]: 'nation',
            source_columns[1]: 'type',
            source_columns[2]: 'model',
        }
    )
    combined = pd.concat([master_frame[['nation', 'type', 'model']], spec_rows], ignore_index=True)
    combined = clean_text_columns(combined, ['nation', 'type', 'model'])
    combined = combined.dropna(subset=['nation', 'type', 'model'])
    combined = combined.drop_duplicates(subset=['nation', 'type', 'model'], keep='first')
    return combined


def prepare_specifications(frame):
    # Keep only usable specification rows and collapse duplicate pairings to a
    # single final record per tanker/receiver combination.
    cleaned = clean_text_columns(
        frame,
        PAIR_COLUMNS + [
            'c_tanker',
            'c_receiver',
            'v_srd_tanker',
            'v_srd_receiver',
            'boom_pod_bda',
            'notes',
        ],
    )
    cleaned = cleaned.dropna(subset=PAIR_COLUMNS)

    duplicate_count = int(cleaned.duplicated(subset=PAIR_COLUMNS).sum())
    if duplicate_count:
        print(
            f"Warning: {duplicate_count} duplicate specification rows found. "
            "Keeping the last row per tanker/receiver combination."
        )

    cleaned = cleaned.drop_duplicates(subset=PAIR_COLUMNS, keep='last')
    return cleaned


# Read Excel sheets by name
tankers_excel = pd.read_excel(excel_path, sheet_name='Tankers')
receivers_excel = pd.read_excel(excel_path, sheet_name='Receivers')
specifications_excel = pd.read_excel(excel_path, sheet_name='Specifications')
# Load this column separately so Excel formulas are preserved as numeric values.
specifications_excel['fuel_flow_rate'] = load_formula_backed_column(
    excel_path,
    'Specifications',
    'fuel_flow_rate',
)

tankers_excel = clean_text_columns(tankers_excel, ['nation', 'type', 'model'])
receivers_excel = clean_text_columns(receivers_excel, ['nation', 'type', 'model'])
specifications_excel = prepare_specifications(specifications_excel)

tankers_final = build_master_rows(
    tankers_excel,
    specifications_excel,
    ['tanker_nation', 'tanker_type', 'tanker_model'],
)
receivers_final = build_master_rows(
    receivers_excel,
    specifications_excel,
    ['receiver_nation', 'receiver_type', 'receiver_model'],
)

with engine.begin() as conn:
    # 1 Ensure base tables exist
    conn.execute(text(create_tables_sql))

    # 2 Clear old data
    conn.execute(text("""
                      TRUNCATE tankers, receivers, compatibility 
                      RESTART IDENTITY CASCADE;
                      """))

    print("Tables cleared")

    # 3 Upload tankers and receivers
    tankers_final.to_sql('tankers', conn, if_exists='append', index=False)
    receivers_final.to_sql('receivers', conn, if_exists='append', index=False)
    print("Tankers and receivers inserted")

    # 4 Read IDs back
    df_tankers = pd.read_sql(
        "SELECT id as tanker_id, nation, type, model FROM tankers", conn
    )
    df_receivers = pd.read_sql(
        "SELECT id as receiver_id, nation, type, model FROM receivers", conn
    )

    # 5 Generate compatibility only from unique combinations found in Specifications
    # Match specification rows back to the inserted tanker and receiver IDs so
    # compatibility only contains real supported pairs.
    compatibility_pairs = specifications_excel[PAIR_COLUMNS].merge(
        df_tankers,
        left_on=['tanker_nation', 'tanker_type', 'tanker_model'],
        right_on=['nation', 'type', 'model'],
        how='inner'
    ).merge(
        df_receivers,
        left_on=['receiver_nation', 'receiver_type', 'receiver_model'],
        right_on=['nation', 'type', 'model'],
        how='inner',
        suffixes=('_tanker', '_receiver')
    )[['tanker_id', 'receiver_id']].drop_duplicates()

    compatibility_pairs.to_sql('compatibility', conn, if_exists='append', index=False)
    print("Compatibility rows inserted from Specifications")

    df_comp = pd.read_sql(
        "SELECT id as compatibility_id, tanker_id, receiver_id FROM compatibility", conn
    )

    # 6 Merge specifications with IDs
    # Resolve every specification row to its compatibility_id before inserting
    # the detailed operational data into the specifications table.
    specs_with_ids = specifications_excel.merge(
        df_tankers,
        left_on=['tanker_nation', 'tanker_type', 'tanker_model'],
        right_on=['nation', 'type', 'model']
    )
    specs_with_ids = specs_with_ids.merge(
        df_receivers,
        left_on=['receiver_nation', 'receiver_type', 'receiver_model'],
        right_on=['nation', 'type', 'model']
    )
    final_specs = specs_with_ids.merge(
        df_comp,
        on=['tanker_id', 'receiver_id']
    )
    columns_to_keep = [
        'compatibility_id',
        'c_tanker',
        'c_receiver',
        'v_srd_tanker',
        'v_srd_receiver',
        'boom_pod_bda',
        'min_alt',
        'max_alt',
        'min_as',
        'max_as_kcas',
        'max_as_m',
        'fuel_flow_rate',
        'notes'
    ]
    final_specs[columns_to_keep].to_sql(
        'specifications',
        conn,
        if_exists='append',
        index=False
    )

print("Excel sheets successfully written to PostgreSQL!")
